import openpyxl as xl
from openpyxl import Workbook, load_workbook, drawing
from openpyxl.utils.cell import get_column_letter
from copy import copy
import pandas as pd
from PIL import Image as PILImage
import itertools
# import xlsxwriter

def convert_file(raw_filename, file_location):
    wb = Workbook()
    ws = wb.active

    file_name = raw_filename
    file_path = file_location
    sheet_name = 'Sheet1' #changes from Sheet1
    # converted_file = file_location

    # file_name = 'Test Case Cartensz - Checkout'
    # file_path = 'file_raw/'+file_name+'.xlsx' # Source file
    # sheet_name = 'Sheet1'
    # converted_file = 'converted/'+file_name+'_openpyxl3.xlsx'

    # wb.save(converted_file)

    converted_wb = xl.load_workbook(filename=file_path)
    converted_ws = converted_wb.create_sheet('Test Report')
    # converted_ws = converted_wb.create_sheet()

    df = pd.read_excel(file_path, sheet_name=sheet_name) # Read excel

    # workbook = xl.Workbook(converted_file)
    # converted_ws = workbook.create_sheet(title="Sheet1")

    # converted_ws = workbook.active
    # converted_ws.title = "Sheet1"

    start_row = 0 #change to 0 to fix first value is missing
    end_row = 250

    # -------------------- HEADER --------------------
    # ---- template read -----
    # template_file = 'template_test_case/Template Test Report Document.xlsx'
    template_file = 'testcase_template/Template Test Report Document FIX.xlsx'
    # converted_file2 = filename + '.xlsx'

    template_wb = xl.load_workbook(filename=template_file)
    # ws1 = template_wb.worksheets[0]
    header = template_wb.worksheets[0]
    footer = template_wb.worksheets[1]
    testcase_format = template_wb.worksheets[2]

    # converted_wb = xl.load_workbook(filename=converted_file2)
    # converted_ws = converted_wb.create_sheet(ws1.title)
    # converted_ws = converted_wb.create_sheet('Test Report')

    # sheet_names = converted_wb.sheetnames # check converted file

    # name_sheet = []

    # # Print the sheet names
    # for sheet_name in sheet_names:
    #     name_sheet.append(sheet_name)
    #     # print(sheet_name)

    # print(name_sheet[0])

    def copy_style(origin_cell, target_cell):
        if origin_cell.has_style:
            target_cell.font = copy(origin_cell.font)
            target_cell.border = copy(origin_cell.border)
            target_cell.fill = copy(origin_cell.fill)
            target_cell.number_format = copy(origin_cell.number_format)
            target_cell.protection = copy(origin_cell.protection)
            target_cell.alignment = copy(origin_cell.alignment)

    def resize_image(image_path, max_width, max_height):
        with PILImage.open(image_path) as img:
            original_width, original_height = img.size
            ratio = min(max_width / original_width, max_height / original_height)
            new_size = (int(original_width * ratio), int(original_height * ratio))
            # resized_img = img.resize(new_size, PILImage.ANTIALIAS)
            resized_img = img.resize(new_size, PILImage.LANCZOS)
            resized_img.save('resized_image.png')
            return 'resized_image.png'

    # ----- HEADER -----
    # Iterate through the columns of the source sheet
    for column_index in range(1, header.max_column + 1):
        # Get the column letter for the current column index
        column_letter = get_column_letter(column_index)

        # Get the max row index of the merged range in the current column
        max_row_index = 0
        for merged_range in header.merged_cells.ranges:
            if merged_range.min_col <= column_index <= merged_range.max_col:
                max_row_index = max(max_row_index, merged_range.max_row)

        # Set the width of the column in the destination sheet
        converted_ws.column_dimensions[column_letter].width = header.column_dimensions[column_letter].width

        # Copy cell values and styles from source to destination
        for row_index in range(1, max_row_index + 1):
            source_cell = header.cell(row=row_index, column=column_index)
            dest_cell = converted_ws.cell(row=row_index, column=column_index)
            dest_cell.value = source_cell.value
            # dest_cell.style = source_cell.style

            # dest_cell.style = NamedStyle(name=source_cell.style) if source_cell.style else None
            converted_ws.row_dimensions[row_index].height = header.row_dimensions[row_index].height # test copy row height
            row_index_header = max_row_index #check index

            if source_cell.has_style:
                dest_cell.font = copy(source_cell.font)
                dest_cell.border = copy(source_cell.border)
                dest_cell.fill = copy(source_cell.fill)
                dest_cell.number_format = copy(source_cell.number_format)
                dest_cell.protection = copy(source_cell.protection)
                dest_cell.alignment = copy(source_cell.alignment)

    # print(row_index_header) #check index

    # ------------------------------------------------

    # Print the column names in the sheet
    # columns_in_sheet = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=1).iloc[0]
    # print(f"Columns in the sheet: {list(columns_in_sheet)}")

    column_name1 = 'TEST OBJECTIVE'
    column_name2 = 'PRECONDITION'
    column_name3 = 'EXPECTED RESULT'
    column_name4 = 'TYPE'
    column_name5 = 'TEST RESULT' #TEST STG
    column_name6 = 'TEST DATA'
    column_name7 = 'FEATURE'

    # ----- INITIATION DATA -----

    # Test Case Form
    test_objective = df.loc[start_row:end_row, column_name1].fillna("")
    precondition = df.loc[start_row:end_row, column_name2].fillna("")
    expected_result = df.loc[start_row:end_row, column_name3].fillna("")
    test_type = df.loc[start_row:end_row, column_name4].fillna("")
    test_result_raw = df.loc[start_row:end_row, column_name5].fillna("")
    steps = df.loc[start_row:end_row, column_name6].fillna("")
    feature = df.loc[start_row:end_row, column_name7].fillna("")

    new_data = list(zip(test_objective, precondition, expected_result, test_type, test_result_raw, steps, feature))
    # print(new_data)

    header = ['No', 'Use Case/Process', '', 'Test Type', 'Smoke/Sanity Test', 'Test Result']#, 'Tester Name', 'Test Date', 'Feedback', 'Evidence', 'Remark']

    # ----- WRITE DATA -----

    row = row_index_header + 1 #0
    column = 2 #0

    # commented for copy head from template, not generate
    # for item in header :
    #     converted_ws.cell(row, column, value=item)
    #     column += 1

    converted_ws.merge_cells('C' + str(row_index_header + 1) + ':D' + str(row_index_header + 1), 'Use Case/Process') # Merge Use Case cell
    converted_ws.merge_cells('B' + str(row_index_header + 2) + ':G' + str(row_index_header + 2))
    converted_ws.merge_cells('B' + str(row_index_header + 3) + ':G' + str(row_index_header + 3))
    for h in range(0,3):
        copy_style(testcase_format['A17'], converted_ws['A' + str(row_index_header + 1 + h)])
        copy_style(testcase_format['H17'], converted_ws['H' + str(row_index_header + 1 + h)])
        copy_style(testcase_format['I17'], converted_ws['I' + str(row_index_header + 1 + h)])

    # img = drawing.image.Image('template_test_case/privy_logo.png')
    # maxsize = (305, 127)
    # img_import = img.thumbnail(maxsize, PIL.Image.ANTIALIAS)
    # converted_ws.add_image(img, 'A1')

    resized_image_path = resize_image('testcase_template/privy_logo.png', 408, 168 ) #412, 172)#305, 127)
    img = drawing.image.Image(resized_image_path)
    converted_ws.add_image(img, 'A1')

    row1 = row_index_header + 4 #1
    column1 = 2 #0
    number = 1

    for a, b, c, d, e, f, g in (new_data) :
        # converted_ws.merge_cells('A'+str(row1+1)+':A'+str(row1+7), value=number) # numbering
        converted_ws['B'+str(row1)].value=number # numbering
        converted_ws.merge_cells('B'+str(row1)+':B'+str(row1+6)) # numbering
        converted_ws.cell(row1, column1 + 1, value="Use Case Uji:")
        converted_ws.cell(row1, column1 + 2, value=g) # case
        converted_ws.cell(row1 + 1, column1 + 1, value="Test Description:")
        converted_ws.cell(row1 + 1, column1 + 2, value=a)
        converted_ws.cell(row1 + 2, column1 + 1, value="Precondition:")
        converted_ws.cell(row1 + 2, column1 + 2, value=b) # precondition
        converted_ws['C'+str(row1+3)].value="Test Case/Test Data:"
        converted_ws.merge_cells('C'+str(row1+3)+':C'+str(row1+5))
        converted_ws['D'+str(row1+3)].value=f
        converted_ws.merge_cells('D'+str(row1+3)+':D'+str(row1+5))
        converted_ws.cell(row1 + 6, column1 + 1, value="Expected Result:")
        converted_ws.cell(row1 + 6, column1 + 2, value=c) # expected result
        converted_ws['E'+str(row1)].value=d # issue in write type
        converted_ws.merge_cells('E'+str(row1)+':E'+str(row1+6)) # issue in write type
        converted_ws['F'+str(row1)].value="" # smoke or sanity
        converted_ws.merge_cells('F'+str(row1)+':F'+str(row1+6)) # smoke or sanity
        converted_ws['G'+str(row1)].value=e # test result
        converted_ws.merge_cells('G'+str(row1)+':G'+str(row1+6)) # test result

        for l in range(0,7):
            copy_style(testcase_format['A17'], converted_ws['A'+str(row1+l)])
        for m in range(0,7):
            copy_style(testcase_format['B17'], converted_ws['B'+str(row1+m)])
            copy_style(testcase_format['C17'], converted_ws['C'+str(row1)])
            copy_style(testcase_format['D17'], converted_ws['D'+str(row1)])
            copy_style(testcase_format['C18'], converted_ws['C'+str(row1+1)])
            copy_style(testcase_format['D18'], converted_ws['D'+str(row1+1)])
            copy_style(testcase_format['C19'], converted_ws['C'+str(row1+2)])
            copy_style(testcase_format['D19'], converted_ws['D'+str(row1+2)])
        for n in range(0,3):
            copy_style(testcase_format['C20'], converted_ws['C'+str(row1+3+n)])
            copy_style(testcase_format['D20'], converted_ws['D'+str(row1+3+n)])
            copy_style(testcase_format['C23'], converted_ws['C'+str(row1+6)])
            copy_style(testcase_format['D23'], converted_ws['D'+str(row1+6)])
        for o in range(0,7):
            copy_style(testcase_format['E17'], converted_ws['E'+str(row1+o)])
            copy_style(testcase_format['F17'], converted_ws['F'+str(row1+o)])
            copy_style(testcase_format['G17'], converted_ws['G'+str(row1+o)])
            copy_style(testcase_format['H17'], converted_ws['H'+str(row1+o)])
            copy_style(testcase_format['I17'], converted_ws['I'+str(row1+o)])


        # copy row height
        converted_ws.row_dimensions[row1].height = testcase_format.row_dimensions[17].height
        converted_ws.row_dimensions[row1+1].height = testcase_format.row_dimensions[18].height
        converted_ws.row_dimensions[row1+2].height = testcase_format.row_dimensions[19].height
        converted_ws.row_dimensions[row1+3].height = testcase_format.row_dimensions[20].height
        converted_ws.row_dimensions[row1+4].height = testcase_format.row_dimensions[21].height
        converted_ws.row_dimensions[row1+5].height = testcase_format.row_dimensions[22].height
        converted_ws.row_dimensions[row1+6].height = testcase_format.row_dimensions[23].height

        number += 1
        row1 += 7
        convert_index = row1

    # print(convert_index)

    # -------------------- FOOTER --------------------
    # ----- FOOTER -----
    for column_index in range(1, footer.max_column + 1):
    # for column_index in range(1, footer.max_column + 14):
        # Get the column letter for the current column index
        column_letter = get_column_letter(column_index)

        # Get the max row index of the merged range in the current column
        max_row_index = 0
        for merged_range in footer.merged_cells.ranges:
            if merged_range.min_col <= column_index <= merged_range.max_col:
                max_row_index = max(max_row_index, merged_range.max_row)

        # Set the width of the column in the destination sheet
        converted_ws.column_dimensions[column_letter].width = footer.column_dimensions[column_letter].width

        # Copy cell values and styles from source to destination
        for row_index in range(1, max_row_index + 1):
            source_cell = footer.cell(row=row_index, column=column_index)
            dest_cell = converted_ws.cell(row=row_index + convert_index, column=column_index)
            # dest_cell = converted_ws.cell(row=row_index, column=column_index)
            dest_cell.value = source_cell.value
            # dest_cell.style = source_cell.style

            # dest_cell.style = NamedStyle(name=source_cell.style) if source_cell.style else None
            converted_ws.row_dimensions[row_index + convert_index].height = footer.row_dimensions[row_index].height # test copy row height
            
            if source_cell.has_style:
                dest_cell.font = copy(source_cell.font)
                dest_cell.border = copy(source_cell.border)
                dest_cell.fill = copy(source_cell.fill)
                dest_cell.number_format = copy(source_cell.number_format)
                dest_cell.protection = copy(source_cell.protection)
                dest_cell.alignment = copy(source_cell.alignment)

    # SETUP MERGE CELLS HEADER
    converted_ws.merge_cells('A1:C5')
    converted_ws.merge_cells('D1:G1')
    converted_ws.merge_cells('D2:G2')
    converted_ws.merge_cells('D3:G3')
    converted_ws.merge_cells('D4:G4')
    converted_ws.merge_cells('D5:G5')
    converted_ws.merge_cells('H5:I5')
    converted_ws.merge_cells('A6:I6')
    converted_ws.merge_cells('B7:D7')
    converted_ws.merge_cells('B8:C8')
    converted_ws.merge_cells('B9:C9')
    converted_ws.merge_cells('B10:C10')
    converted_ws.merge_cells('B11:C11')
    converted_ws.merge_cells('B12:C12')
    converted_ws.merge_cells('E7:I12')
    converted_ws.merge_cells('A13:I13')

    # SETUP MERGE CELLS FOOTER
    footer_num = convert_index #16
    converted_ws.merge_cells('B'+str(1+footer_num)+':I'+str(1+footer_num))
    converted_ws.merge_cells('B'+str(2+footer_num)+':I'+str(2+footer_num))
    converted_ws.merge_cells('A'+str(9+footer_num)+':I'+str(9+footer_num))
    # converted_ws.delete_rows(10+footer_num)
    converted_ws.merge_cells('B'+str(11+footer_num)+':F'+str(11+footer_num))
    # converted_ws.merge_cells('E'+str(11+footer_num)+':F'+str(11+footer_num)) # sign

    for foot1 in range(11,20) : 
        converted_ws.merge_cells('E'+str(foot1+footer_num)+':F'+str(foot1+footer_num))

    # converted_ws.merge_cells('B'+str(11+footer_num)+':F'+str(11+footer_num))

    # bottom sign
    converted_ws.merge_cells('A'+str(20+footer_num)+':I'+str(20+footer_num)) # separator

    for foot2 in range(21,25) : 
        converted_ws.merge_cells('B'+str(foot2+footer_num)+':C'+str(foot2+footer_num)) # prepared by

    for foot3 in range(21,25) : 
        converted_ws.merge_cells('E'+str(foot3+footer_num)+':F'+str(foot3+footer_num)) # reviewed by

    # ----- hardcoded for approved by -----
    # converted_ws.delete_rows(footer_num + 21)
    converted_ws['D' + str(footer_num + 21 + 1)] = 'Approved by:'
    converted_ws['D' + str(footer_num + 23 + 1)] = '[Nama]'
    converted_ws['D' + str(footer_num + 24 + 1)] = 'Product Manager'

    converted_ws.merge_cells('B' + str(footer_num + 24 + 1) + ':C' + str(footer_num + 24 + 1)) # technical writer
    converted_ws.merge_cells('E' + str(footer_num + 24 + 1) + ':F' + str(footer_num + 24 + 1)) # vice president of engineering

    for x in range(4):
        source_cl = footer['D'+str(22 + x)]
        dest_cl = converted_ws['D'+str(footer_num + 21 + 1 + x)]

        if source_cl.has_style:
            dest_cl.font = copy(source_cl.font)
            dest_cl.border = copy(source_cl.border)
            dest_cl.fill = copy(source_cl.fill)
            dest_cl.number_format = copy(source_cl.number_format)
            dest_cl.protection = copy(source_cl.protection)
            dest_cl.alignment = copy(source_cl.alignment)

    # align G column 
    converted_ws.column_dimensions['G'].width = testcase_format.column_dimensions['F'].width
    # ------------------------------------------------

    # ----- cleaning -----
    # converted_ws.delete_rows(10+footer_num)
    # converted_ws.delete_rows(21+footer_num)

    # workbook.close()
    converted_wb.save(file_path)

# ----- Created by FK -----